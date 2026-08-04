"""Production seed: Excel ACTIVE campaigns → campaigns table (Postgres).

Schema uyumu (campaign/models.py + PostgresCampaignRepository):
  campaigns(
    id, campaign_code, title, summary, category_id,
    brand, product_name, list_price, currency,
    installment_count, monthly_payment, cash_price,
    min_budget, max_budget, membership_required,
    membership_cta_url, membership_cta_label,
    starts_at, ends_at, status, attributes, search_text, updated_at
  )
  categories(id, category_code, ...)

Usage:
  python -m taksitlio.campaign.seed_from_excel \\
    --excel path/to/Kategoriler_ve_Kampanya_Ornekleri.xlsx \\
    --database-url $DATABASE_URL \\
    [--category-code 1] \\
    [--dry-run] [-v]
"""

from __future__ import annotations

import argparse
import asyncio
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

import openpyxl

from taksitlio.campaign.seed_budget_fix import infer_budget_bounds

logger = logging.getLogger(__name__)

# Excel ACTIVE rows that are general retail finance — apply to phone (1) by default
# when no explicit category is present in the feed.
DEFAULT_CATEGORY_CODE = "1"  # Cep Telefonu


def _parse_excel_date(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value
    return None


def _extract_rate_text(text: str) -> Optional[str]:
    if not text:
        return None
    patterns = [
        r"%\s*[\d,\.]+\s*(?:kar\s*oranı|faiz|kar\s*payı)",
        r"%\s*0\s*(?:faizli|oran)",
        r"aylık\s*(?:kar\s*)?oran[ıi]\s*%\s*[\d,\.]+",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return m.group(0).strip()
    return None


def _infer_bank(title: str, text: str) -> Optional[str]:
    combined = f"{title} {text}".lower()
    banks = {
        "albaraka": "Albaraka Türk",
        "kuveyt": "Kuveyt Türk",
        "getirfinans": "Getir Finans",
        "getir finans": "Getir Finans",
        "fibabanka": "Fibabanka",
        "qnb": "QNB Finansbank",
        "yapı kredi": "Yapı Kredi",
        "iş bankası": "İş Bankası",
        "akbank": "Akbank",
        "garanti": "Garanti BBVA",
    }
    for key, name in banks.items():
        if key in combined:
            return name
    return None


def _infer_tenure(text: str) -> Optional[int]:
    m = re.search(r"(\d+)\s*ay", text or "", re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def load_active_campaigns(excel_path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "Kampanyalar" not in wb.sheetnames:
        raise ValueError("Sheet 'Kampanyalar' not found")

    ws = wb["Kampanyalar"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    campaigns: list[dict[str, Any]] = []

    for raw in rows[1:]:
        row = dict(zip(headers, raw))
        status = (row.get("status") or "").upper().strip()
        if status != "ACTIVE":
            continue

        external_id = row.get("id")
        title = (row.get("title") or "").strip()
        text = (row.get("text") or "").strip()
        subtitle = (row.get("subtitle") or "").strip() or None
        bank = _infer_bank(title, text)
        rate_text = _extract_rate_text(text)
        min_b, max_b = infer_budget_bounds(text, subtitle or "")
        min_budget = min_b if min_b is not None else 1000.0
        max_budget = max_b  # None = no ceiling → eligibility unrestricted
        tenure = _infer_tenure(text)

        campaign_code = f"EXCEL-{external_id}" if external_id is not None else f"EXCEL-{hash(title) % 10_000_000}"

        camp = {
            "external_id": int(external_id) if external_id is not None else None,
            "campaign_code": campaign_code,
            "title": title,
            "summary": subtitle or title,
            "text": text,
            "status": "ACTIVE",
            "type": (row.get("type") or "RETAIL").upper(),
            "starts_at": _parse_excel_date(row.get("activity_start_date")),
            "ends_at": _parse_excel_date(row.get("activity_end_date")),
            "bank": bank,
            "brand": bank,
            "rate_text": rate_text,
            "max_budget": max_budget,
            "min_budget": min_budget,
            "installment_count": tenure,
            "membership_required": True,
            "membership_cta_label": "Üye ol, kampanyadan yararlan",
            "membership_cta_url": None,
            "currency": "TRY",
            "attributes": {
                "source": "excel_seed",
                "excel_id": external_id,
                "rate_text": rate_text,
                "raw_subtitle": subtitle,
            },
            "search_text": f"{title} {subtitle or ''} {bank or ''} {rate_text or ''}".strip(),
        }
        campaigns.append(camp)

    logger.info("Loaded %d ACTIVE campaigns from %s", len(campaigns), excel_path)
    return campaigns


async def _resolve_category_id(conn: Any, category_code: str) -> int:
    raw = str(category_code).strip()
    # Fehmi Excel numeric ids (1=Cep Telefonu) → taxonomy codes used in prod DB.
    excel_id_aliases = {
        "1": "MOBILE_PHONE",
        "3": "TABLET",
        "4": "LAPTOP",
    }
    candidates = [raw]
    if raw in excel_id_aliases:
        candidates.append(excel_id_aliases[raw])
    if not raw.isdigit():
        candidates.append(raw.upper())

    for code in candidates:
        row = await conn.fetchrow(
            """
            SELECT id FROM categories
            WHERE category_code = $1 AND status = 'ACTIVE'
            LIMIT 1
            """,
            code,
        )
        if row is None:
            row = await conn.fetchrow(
                "SELECT id FROM categories WHERE category_code = $1 LIMIT 1",
                code,
            )
        if row is not None:
            return int(row["id"])

    # Last resort: treat numeric input as categories.id (prod: id=1 → MOBILE_PHONE).
    if raw.isdigit():
        row = await conn.fetchrow(
            "SELECT id FROM categories WHERE id = $1 LIMIT 1",
            int(raw),
        )
        if row is not None:
            return int(row["id"])

    raise RuntimeError(
        f"Category code '{category_code}' not found in categories table. "
        "Seed categories first or pass --category-code that exists "
        "(e.g. MOBILE_PHONE, or Excel id 1)."
    )


async def upsert_campaigns(
    campaigns: list[dict[str, Any]],
    *,
    database_url: str,
    category_code: str = DEFAULT_CATEGORY_CODE,
    dry_run: bool = False,
) -> int:
    if dry_run:
        for c in campaigns:
            logger.info(
                "[DRY-RUN] upsert code=%s title=%r bank=%s rate=%s max_budget=%s",
                c["campaign_code"],
                c["title"][:55],
                c.get("bank"),
                c.get("rate_text"),
                c.get("max_budget"),
            )
        return len(campaigns)

    import asyncpg

    pool = await asyncpg.create_pool(database_url, min_size=1, max_size=3)
    assert pool is not None
    count = 0
    try:
        async with pool.acquire() as conn:
            category_id = await _resolve_category_id(conn, category_code)
            logger.info("Resolved category_code=%s → category_id=%s", category_code, category_id)

            for c in campaigns:
                await conn.execute(
                    """
                    INSERT INTO campaigns (
                        campaign_code, title, summary, category_id,
                        brand, product_name, list_price, currency,
                        installment_count, monthly_payment, cash_price,
                        min_budget, max_budget,
                        membership_required, membership_cta_url, membership_cta_label,
                        starts_at, ends_at, status, attributes, search_text,
                        updated_at
                    ) VALUES (
                        $1, $2, $3, $4,
                        $5, $6, $7, $8,
                        $9, $10, $11,
                        $12, $13,
                        $14, $15, $16,
                        $17, $18, $19, $20::jsonb, $21,
                        NOW()
                    )
                    ON CONFLICT (campaign_code) DO UPDATE SET
                        title = EXCLUDED.title,
                        summary = EXCLUDED.summary,
                        category_id = EXCLUDED.category_id,
                        brand = EXCLUDED.brand,
                        installment_count = EXCLUDED.installment_count,
                        min_budget = EXCLUDED.min_budget,
                        max_budget = EXCLUDED.max_budget,
                        membership_required = EXCLUDED.membership_required,
                        membership_cta_label = EXCLUDED.membership_cta_label,
                        starts_at = EXCLUDED.starts_at,
                        ends_at = EXCLUDED.ends_at,
                        status = EXCLUDED.status,
                        attributes = EXCLUDED.attributes,
                        search_text = EXCLUDED.search_text,
                        updated_at = NOW()
                    """,
                    c["campaign_code"],
                    c["title"],
                    c["summary"],
                    category_id,
                    c.get("brand"),
                    None,  # product_name
                    None,  # list_price
                    c.get("currency") or "TRY",
                    c.get("installment_count"),
                    None,  # monthly_payment
                    None,  # cash_price
                    c.get("min_budget"),
                    c.get("max_budget"),
                    bool(c.get("membership_required", True)),
                    c.get("membership_cta_url"),
                    c.get("membership_cta_label"),
                    c.get("starts_at"),
                    c.get("ends_at"),
                    c.get("status") or "ACTIVE",
                    __import__("json").dumps(c.get("attributes") or {}),
                    c.get("search_text") or "",
                )
                count += 1
                logger.info("Upserted %s (%s)", c["campaign_code"], c.get("bank"))
    finally:
        await pool.close()
    return count


def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Seed ACTIVE campaigns from Excel → campaigns table")
    parser.add_argument("--excel", required=True, type=Path)
    parser.add_argument("--database-url", default=None)
    parser.add_argument(
        "--category-code",
        default=DEFAULT_CATEGORY_CODE,
        help="Category code to attach (default: 1 = Cep Telefonu)",
    )
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    if not args.excel.exists():
        logger.error("Excel not found: %s", args.excel)
        return 1

    campaigns = load_active_campaigns(args.excel)
    if not campaigns:
        logger.warning("No ACTIVE campaigns found")
        return 0

    if args.dry_run or not args.database_url:
        asyncio.run(
            upsert_campaigns(
                campaigns,
                database_url="",
                category_code=args.category_code,
                dry_run=True,
            )
        )
        return 0

    n = asyncio.run(
        upsert_campaigns(
            campaigns,
            database_url=args.database_url,
            category_code=args.category_code,
            dry_run=False,
        )
    )
    logger.info("Upserted %d campaigns into category_code=%s", n, args.category_code)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
