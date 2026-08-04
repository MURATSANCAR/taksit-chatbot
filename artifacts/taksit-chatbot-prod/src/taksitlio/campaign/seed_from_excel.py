"""Production seed script: load ACTIVE campaigns from the product Excel
into the campaign catalog (finance_campaigns / legacy campaigns table).

Usage (from repo root):
    python -m taksitlio.campaign.seed_from_excel \
        --excel path/to/Kategoriler_ve_Kampanya_Ornekleri.xlsx \
        --database-url $DATABASE_URL \
        --dry-run

Idempotent: campaigns are upserted by external_id (Excel id column).
Only rows with status=ACTIVE are inserted/updated.
"""

from __future__ import annotations

import argparse
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

import openpyxl

logger = logging.getLogger(__name__)


def _parse_excel_date(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value
    return None


def _extract_rate_text(text: str) -> Optional[str]:
    """Pull a human-readable rate snippet from the long campaign text."""
    if not text:
        return None
    # Common patterns in the sample data
    patterns = [
        r"%\s*[\d,\.]+\s*(?:kar\s*oranı|faiz|kar\s*payı)",
        r"%\s*0\s*(?:faizli|oran)",
        r"aylık\s*(?:kar\s*)?oran[ıi]\s*%\s*[\d,\.]+",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return m.group(0).strip()
    return None


def _infer_bank(title: str, text: str) -> Optional[str]:
    combined = f"{title} {text}".lower()
    banks = {
        "albaraka": "Albaraka Türk",
        "kuveyt": "Kuveyt Türk",
        "getirfinans": "Getir Finans",
        "getir finans": "Getir Finans",
        "fibabanka": "Fibabanka",
        "qnb": "QNB Finansbank",
        "yapı kredi": "Yapı Kredi",
        "iş bankası": "İş Bankası",
        "akbank": "Akbank",
        "garanti": "Garanti BBVA",
    }
    for key, name in banks.items():
        if key in combined:
            return name
    return None


def load_active_campaigns(excel_path: Path) -> list[dict[str, Any]]:
    """Parse the Kampanyalar sheet and return only ACTIVE rows as dicts."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "Kampanyalar" not in wb.sheetnames:
        raise ValueError("Sheet 'Kampanyalar' not found")

    ws = wb["Kampanyalar"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    campaigns: list[dict[str, Any]] = []

    for raw in rows[1:]:
        row = dict(zip(headers, raw))
        status = (row.get("status") or "").upper().strip()
        if status != "ACTIVE":
            continue

        external_id = row.get("id")
        title = (row.get("title") or "").strip()
        text = (row.get("text") or "").strip()
        subtitle = (row.get("subtitle") or "").strip() or None

        camp = {
            "external_id": int(external_id) if external_id is not None else None,
            "title": title,
            "subtitle": subtitle,
            "text": text,
            "status": "ACTIVE",
            "type": (row.get("type") or "RETAIL").upper(),
            "activity_start_date": _parse_excel_date(row.get("activity_start_date")),
            "activity_end_date": _parse_excel_date(row.get("activity_end_date")),
            "sequence": row.get("sequence"),
            "last_updated": _parse_excel_date(row.get("last_updated")),
            "bank": _infer_bank(title, text),
            "rate_text": _extract_rate_text(text),
            "summary": subtitle or title,
            # Category linkage is left to the ranking / eligibility layer
            # (sample data does not carry explicit category_id)
            "category_codes": None,
            "raw": {k: v for k, v in row.items() if v is not None},
        }
        campaigns.append(camp)

    logger.info("Loaded %d ACTIVE campaigns from %s", len(campaigns), excel_path)
    return campaigns


def upsert_campaigns(
    campaigns: list[dict[str, Any]],
    *,
    database_url: str,
    dry_run: bool = False,
) -> int:
    """
    Upsert into the production campaign table.

    NOTE: The exact table name / schema follows ADR-010 (finance_campaigns).
    This function is intentionally written against a minimal, stable interface
    so it can be adapted when the final schema is locked.
    """
    if dry_run:
        for c in campaigns:
            logger.info(
                "[DRY-RUN] would upsert id=%s title=%r bank=%s rate=%s",
                c["external_id"],
                c["title"][:60],
                c.get("bank"),
                c.get("rate_text"),
            )
        return len(campaigns)

    # Production path – uses the project's own repository / SQLAlchemy models.
    # Import is local so the seed script can still be imported in isolation.
    try:
        from taksitlio.campaign.repository import CampaignRepository  # type: ignore
        from taksitlio.db.session import get_session  # type: ignore
    except ImportError:
        logger.error(
            "Project DB layer not importable. "
            "Run this script from the repository root with the package installed."
        )
        raise

    count = 0
    with get_session(database_url) as session:
        repo = CampaignRepository(session)
        for camp in campaigns:
            repo.upsert_from_seed(camp)
            count += 1
        session.commit()
    return count


def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Seed ACTIVE campaigns from Excel")
    parser.add_argument("--excel", required=True, type=Path, help="Path to the Excel file")
    parser.add_argument("--database-url", default=None, help="Postgres connection string")
    parser.add_argument("--dry-run", action="store_true", help="Only print what would be done")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    if not args.excel.exists():
        logger.error("Excel file not found: %s", args.excel)
        return 1

    campaigns = load_active_campaigns(args.excel)
    if not campaigns:
        logger.warning("No ACTIVE campaigns found")
        return 0

    if args.dry_run or not args.database_url:
        upsert_campaigns(campaigns, database_url="", dry_run=True)
        return 0

    n = upsert_campaigns(campaigns, database_url=args.database_url, dry_run=False)
    logger.info("Upserted %d campaigns", n)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
